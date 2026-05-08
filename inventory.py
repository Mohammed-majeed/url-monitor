"""
inventory.py
============
Reads the URL_monitoring Excel sheet and produces a list of CheckTarget objects.

A single Excel row may produce multiple CheckTargets — one per filled URL slot
(external_url, internal_url, ingress).

Also recomputes `location_type` based on what's actually filled:
    external + internal (±ingress)        -> "Both"
    external + ingress (no internal)      -> "External"
    internal + ingress (no external)      -> "Internal"
    external only                         -> "External"
    internal only                         -> "Internal"
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

import openpyxl


SHEET_NAME = "URL_monitoring"


@dataclass
class CheckTarget:
    """One URL to be checked."""
    url_id: str                # e.g. URL-ACC-WKP-001
    target_kind: str           # "external" | "internal" | "ingress"
    environment: str           # ONT / TST / ACC / PRD
    location_type: str         # External / Internal / Both  (per-row, recomputed)
    app_name: str
    url: str                   # the URL to actually check
    use_proxy: bool
    runner: str                # ExternalRunner / InternalRunner / Both
    read_body: bool
    expected_text: List[str]   # parsed comma-separated, all must appear
    max_body_bytes: int
    expected_statuses: str     # raw spec, e.g. "200-399;401-login"
    timeout_seconds: int
    notes: str = ""

    @property
    def is_external_runner(self) -> bool:
        return (self.runner or "").lower() == "externalrunner"


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_bool(v, default=False) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().upper()
    return s in ("TRUE", "1", "YES", "Y")


def _parse_expected_text(raw) -> List[str]:
    """Comma-separated list, all required (AND)."""
    s = _norm(raw)
    if not s:
        return []
    return [t.strip() for t in s.split(",") if t.strip()]


def _recompute_location_type(ext: str, intern: str, ingress: str) -> str:
    has_ext = bool(ext)
    has_int = bool(intern)
    has_ing = bool(ingress)

    if has_ext and has_int:
        return "Both"
    if has_ext and has_ing:
        return "External"
    if has_int and has_ing:
        return "Internal"
    if has_ext:
        return "External"
    if has_int:
        return "Internal"
    if has_ing:
        return "Internal"
    return ""


def load_inventory(excel_path: str | os.PathLike) -> List[CheckTarget]:
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {path}")
    ws = wb[SHEET_NAME]

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    targets: List[CheckTarget] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue

        def get(name, default=None):
            idx = col.get(name)
            return row[idx] if idx is not None else default

        url_id = _norm(get("url_id"))
        if not url_id:
            continue

        if not _to_bool(get("enabled"), default=True):
            continue  # skip disabled rows entirely

        environment = _norm(get("environment")).upper()
        app_name = _norm(get("app_name"))

        ext_url = _norm(get("external_url"))
        int_url = _norm(get("internal_url"))
        ing_url = _norm(get("ingress"))

        # Recompute location_type from what's actually present
        loc_type = _recompute_location_type(ext_url, int_url, ing_url)

        expected_text = _parse_expected_text(get("expected_text"))
        max_body = int(get("max_body_bytes") or 200000)
        timeout = int(get("timeout_seconds") or 10)
        expected_statuses = _norm(get("expected_statuses")) or "200-399;401-login"
        read_body = _to_bool(get("read_body"), default=True)
        notes = _norm(get("notes"))

        # Emit one CheckTarget per filled URL slot
        if ext_url:
            targets.append(CheckTarget(
                url_id=url_id,
                target_kind="external",
                environment=environment,
                location_type=loc_type,
                app_name=app_name,
                url=ext_url,
                use_proxy=_to_bool(get("external_proxy"), default=True),
                runner=_norm(get("external_check_from")) or "ExternalRunner",
                read_body=read_body,
                expected_text=expected_text,
                max_body_bytes=max_body,
                expected_statuses=expected_statuses,
                timeout_seconds=timeout,
                notes=notes,
            ))

        if int_url:
            targets.append(CheckTarget(
                url_id=url_id,
                target_kind="internal",
                environment=environment,
                location_type=loc_type,
                app_name=app_name,
                url=int_url,
                use_proxy=_to_bool(get("internal_proxy"), default=False),
                runner=_norm(get("internal_url_from")) or "InternalRunner",
                read_body=read_body,
                expected_text=expected_text,
                max_body_bytes=max_body,
                expected_statuses=expected_statuses,
                timeout_seconds=timeout,
                notes=notes,
            ))

        if ing_url:
            targets.append(CheckTarget(
                url_id=url_id,
                target_kind="ingress",
                environment=environment,
                location_type=loc_type,
                app_name=app_name,
                url=ing_url,
                use_proxy=_to_bool(get("ingress_proxy"), default=False),
                runner=_norm(get("ingress_from")) or "InternalRunner",
                read_body=read_body,
                expected_text=expected_text,
                max_body_bytes=max_body,
                expected_statuses=expected_statuses,
                timeout_seconds=timeout,
                notes=notes,
            ))

    return targets


def split_by_runner(targets: List[CheckTarget]):
    """Split targets into (internal_runner, external_runner) lists."""
    internal_runner: List[CheckTarget] = []
    external_runner: List[CheckTarget] = []
    for t in targets:
        if t.is_external_runner:
            external_runner.append(t)
        else:
            internal_runner.append(t)
    return internal_runner, external_runner


def fix_location_types_in_excel(excel_path: str | os.PathLike) -> int:
    """
    Open the inventory and rewrite the location_type column based on the
    actual content rules. Returns number of rows changed.
    """
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_NAME]

    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers) if h is not None}  # 1-indexed

    changed = 0
    for r in range(2, ws.max_row + 1):
        ext = _norm(ws.cell(row=r, column=col["external_url"]).value)
        intern = _norm(ws.cell(row=r, column=col["internal_url"]).value)
        ing = _norm(ws.cell(row=r, column=col["ingress"]).value)
        new_loc = _recompute_location_type(ext, intern, ing)
        if not new_loc:
            continue
        cell = ws.cell(row=r, column=col["location_type"])
        if _norm(cell.value) != new_loc:
            cell.value = new_loc
            changed += 1

    if changed:
        wb.save(path)
    return changed
