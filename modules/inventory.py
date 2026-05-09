"""
inventory.py
============
Reads the URL_monitoring Excel sheet and produces a list of CheckTarget objects.

A single Excel row may produce multiple CheckTargets — one per enabled and
filled URL slot:
    - enable_ext_url_check     + external_url
    - enable_int_url_check     + internal_url
    - enable_ingress_url_check + ingress

`location_type` is no longer read from or written to Excel. It is derived at
runtime for backward-compatible logging only.

Most defaults are supplied by url_monitor_config.ini through InventoryDefaults.
Excel row values still override these defaults where the matching column exists.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import List

import openpyxl


SHEET_NAME = "URL_monitoring"


@dataclass
class InventoryDefaults:
    """Runtime defaults used when a row/cell is empty or a column is missing."""
    expected_statuses: str = "200-399;401-login"
    read_body: bool = True
    max_body_bytes: int = 200000
    timeout_seconds: int = 10

    external_proxy: bool = True
    internal_proxy: bool = False
    ingress_proxy: bool = False

    external_runner: str = "ExternalRunner"
    internal_runner: str = "InternalRunner"
    ingress_runner: str = "InternalRunner"

    # Backward/forward compatibility behaviour for enable flags.
    legacy_enabled_default: bool = True
    missing_enable_column_default: bool = True
    blank_enable_value_default: bool = False


@dataclass
class CheckTarget:
    """One URL to be checked."""
    url_id: str                # e.g. URL-ACC-WKP-001
    target_kind: str           # "external" | "internal" | "ingress"
    environment: str           # ONT / TST / ACC / PRD
    location_type: str         # External / Internal / Both (derived, not from Excel)
    app_name: str
    url: str                   # the URL to actually check
    use_proxy: bool
    runner: str                # ExternalRunner / InternalRunner / Both
    read_body: bool
    expected_text: List[str]   # parsed comma-separated, all must appear
    max_body_bytes: int
    expected_statuses: str     # raw spec, e.g. "200-399;401-login"
    timeout_seconds: int
    notes: str = ""

    @property
    def is_external_runner(self) -> bool:
        return (self.runner or "").lower() == "externalrunner"


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_bool(v, default=False) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().upper()
    if not s:
        return default
    if s in ("TRUE", "1", "YES", "Y", "JA", "J", "ON"):
        return True
    if s in ("FALSE", "0", "NO", "N", "NEE", "OFF"):
        return False
    return default


def _to_int(v, default: int) -> int:
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(str(v).strip()))
        except (TypeError, ValueError):
            return default


def _parse_expected_text(raw) -> List[str]:
    """Comma-separated list, all required (AND)."""
    s = _norm(raw)
    if not s:
        return []
    return [t.strip() for t in s.split(",") if t.strip()]


def _derive_location_type(ext: str, intern: str, ingress: str) -> str:
    """
    Derive the old log field from the URLs that exist in the row.

    This is kept only so existing Splunk dashboards that expect location_type
    keep working after the Excel column is removed.
    """
    has_ext = bool(ext)
    has_int = bool(intern)
    has_ing = bool(ingress)

    if has_ext and has_int:
        return "Both"
    if has_ext and has_ing:
        return "External"
    if has_int and has_ing:
        return "Internal"
    if has_ext:
        return "External"
    if has_int:
        return "Internal"
    if has_ing:
        return "Internal"
    return ""


def load_inventory(
    excel_path: str | os.PathLike,
    *,
    sheet_name: str = SHEET_NAME,
    defaults: InventoryDefaults | None = None,
) -> List[CheckTarget]:
    """Load enabled URL checks from Excel."""
    defaults = defaults or InventoryDefaults()
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    col = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    def has_col(name: str) -> bool:
        return name in col

    targets: List[CheckTarget] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue

        def get(name, default=None):
            idx = col.get(name)
            return row[idx] if idx is not None and idx < len(row) else default

        url_id = _norm(get("url_id"))
        if not url_id:
            continue

        # Backward compatibility: old files may still have a global enabled column.
        legacy_enabled = (
            _to_bool(get("enabled"), default=defaults.legacy_enabled_default)
            if has_col("enabled")
            else defaults.legacy_enabled_default
        )

        def slot_enabled(enable_column: str) -> bool:
            if has_col(enable_column):
                return legacy_enabled and _to_bool(
                    get(enable_column),
                    default=defaults.blank_enable_value_default,
                )
            return legacy_enabled and defaults.missing_enable_column_default

        environment = _norm(get("environment")).upper()
        app_name = _norm(get("app_name"))

        ext_url = _norm(get("external_url"))
        int_url = _norm(get("internal_url"))
        ing_url = _norm(get("ingress"))

        # Derived only for logs; not read from or written to Excel anymore.
        loc_type = _derive_location_type(ext_url, int_url, ing_url)

        expected_text = _parse_expected_text(get("expected_text"))
        max_body = _to_int(get("max_body_bytes"), defaults.max_body_bytes)
        timeout = _to_int(get("timeout_seconds"), defaults.timeout_seconds)
        expected_statuses = _norm(get("expected_statuses")) or defaults.expected_statuses
        read_body = _to_bool(get("read_body"), default=defaults.read_body)
        notes = _norm(get("notes"))

        # Emit one CheckTarget per enabled + filled URL slot.
        if ext_url and slot_enabled("enable_ext_url_check"):
            targets.append(CheckTarget(
                url_id=url_id,
                target_kind="external",
                environment=environment,
                location_type=loc_type,
                app_name=app_name,
                url=ext_url,
                use_proxy=_to_bool(get("external_proxy"), default=defaults.external_proxy),
                runner=_norm(get("external_check_from")) or defaults.external_runner,
                read_body=read_body,
                expected_text=expected_text,
                max_body_bytes=max_body,
                expected_statuses=expected_statuses,
                timeout_seconds=timeout,
                notes=notes,
            ))

        if int_url and slot_enabled("enable_int_url_check"):
            targets.append(CheckTarget(
                url_id=url_id,
                target_kind="internal",
                environment=environment,
                location_type=loc_type,
                app_name=app_name,
                url=int_url,
                use_proxy=_to_bool(get("internal_proxy"), default=defaults.internal_proxy),
                runner=_norm(get("internal_url_from")) or defaults.internal_runner,
                read_body=read_body,
                expected_text=expected_text,
                max_body_bytes=max_body,
                expected_statuses=expected_statuses,
                timeout_seconds=timeout,
                notes=notes,
            ))

        if ing_url and slot_enabled("enable_ingress_url_check"):
            targets.append(CheckTarget(
                url_id=url_id,
                target_kind="ingress",
                environment=environment,
                location_type=loc_type,
                app_name=app_name,
                url=ing_url,
                use_proxy=_to_bool(get("ingress_proxy"), default=defaults.ingress_proxy),
                runner=_norm(get("ingress_from")) or defaults.ingress_runner,
                read_body=read_body,
                expected_text=expected_text,
                max_body_bytes=max_body,
                expected_statuses=expected_statuses,
                timeout_seconds=timeout,
                notes=notes,
            ))

    return targets


def split_by_runner(targets: List[CheckTarget]):
    """Split targets into (internal_runner, external_runner) lists."""
    internal_runner: List[CheckTarget] = []
    external_runner: List[CheckTarget] = []
    for t in targets:
        if t.is_external_runner:
            external_runner.append(t)
        else:
            internal_runner.append(t)
    return internal_runner, external_runner


def fix_location_types_in_excel(excel_path: str | os.PathLike) -> int:
    """
    Deprecated compatibility shim.

    `location_type` has been removed from the Excel inventory. The value is now
    derived in memory inside `load_inventory()` for logging only. This function
    intentionally does nothing so older imports do not break.
    """
    return 0
